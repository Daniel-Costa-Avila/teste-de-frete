from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


def generate_xlsx(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Produtos"

    # Layout constants
    title_row = 1
    info_row = 2
    header_row = 3
    first_data_row = 4

    headers = ["Grupo", "Nome do produto", "ID do produto", "Link do produto", "CEPs para testar"]
    examples = [
        [
            "Grupo A (Exemplo)",
            "Colchão Casal Premium (Exemplo)",
            "PRD-000001",
            "https://exemplo.com.br/produtos/colchao-casal-premium",
            "01001-000, 20040-002",
        ],
        [
            "Grupo A (Exemplo)",
            "Travesseiro Ortopédico (Exemplo)",
            "PRD-000002",
            "https://exemplo.com.br/produtos/travesseiro-ortopedico",
            "",
        ],
        [
            "Grupo B (Exemplo)",
            "Cama Box Solteiro (Exemplo)",
            "PRD-000003",
            "https://exemplo.com.br/produtos/cama-box-solteiro",
            "79800-002",
        ],
        [
            "Grupo B (Exemplo)",
            "Protetor de Colchão Queen (Exemplo)",
            "PRD-000004",
            "https://exemplo.com.br/produtos/protetor-colchao-queen",
            "04094-050; 30110-012",
        ],
    ]

    # Styles
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    title_font = Font(bold=True, size=14, color="1F2937")
    info_font = Font(size=10, color="374151")

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")  # dark blue
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    body_alignment = Alignment(vertical="top", wrap_text=True)

    warn_fill = PatternFill("solid", fgColor="FDE68A")  # soft yellow
    error_fill = PatternFill("solid", fgColor="FECACA")  # soft red

    # Title + info
    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=len(headers))
    ws.cell(row=title_row, column=1, value="Template de Produtos (Grupos + CEPs)").font = title_font
    ws.cell(row=title_row, column=1).alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(start_row=info_row, start_column=1, end_row=info_row, end_column=len(headers))
    ws.cell(
        row=info_row,
        column=1,
        value="Preencha Link. Opcional: Grupo e CEPs (se vazio, usa o CEP padrão informado na UI). O ID é para controle interno.",
    ).font = info_font
    ws.cell(row=info_row, column=1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.row_dimensions[title_row].height = 22
    ws.row_dimensions[info_row].height = 28

    # Header row
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    ws.row_dimensions[header_row].height = 22

    # Example rows
    for i, row_values in enumerate(examples):
        r = first_data_row + i
        for c, value in enumerate(row_values, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.alignment = body_alignment
            cell.border = border

        link_cell = ws.cell(row=r, column=4)
        if isinstance(link_cell.value, str) and link_cell.value.startswith(("http://", "https://")):
            link_cell.hyperlink = link_cell.value
            link_cell.style = "Hyperlink"

    # Column widths
    ws.column_dimensions["A"].width = 22  # group
    ws.column_dimensions["B"].width = 44  # name
    ws.column_dimensions["C"].width = 18  # id
    ws.column_dimensions["D"].width = 70  # url
    ws.column_dimensions["E"].width = 28  # ceps

    # Freeze panes (keep title+info+header visible)
    ws.freeze_panes = f"A{first_data_row}"

    # Table (includes examples, user can keep adding rows)
    last_row = first_data_row + len(examples) - 1
    table_ref = f"A{header_row}:E{last_row}"
    table = Table(displayName="TabelaProdutos", ref=table_ref)
    table_style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = table_style
    ws.add_table(table)

    # Data validations (apply to a reasonable input range)
    max_rows = 1000
    id_range = f"C{first_data_row}:C{max_rows}"
    link_range = f"D{first_data_row}:D{max_rows}"

    dv_id = DataValidation(
        type="custom",
        formula1=f'=AND(LEN(C{first_data_row})=10,LEFT(C{first_data_row},4)="PRD-",ISNUMBER(VALUE(RIGHT(C{first_data_row},6))))',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="ID inválido",
        error="Use o formato PRD-000001 (PRD- + 6 dígitos).",
    )
    ws.add_data_validation(dv_id)
    dv_id.add(id_range)

    dv_link = DataValidation(
        type="custom",
        formula1=f'=OR(LEFT(D{first_data_row},7)="http://",LEFT(D{first_data_row},8)="https://")',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Link inválido",
        error="Informe um link começando com http:// ou https://",
    )
    ws.add_data_validation(dv_link)
    dv_link.add(link_range)

    # Conditional formatting: highlight missing required fields (Link)
    link_col = "D"
    ws.conditional_formatting.add(
        f"{link_col}{first_data_row}:{link_col}{max_rows}",
        FormulaRule(
            formula=[f'LEN(TRIM({link_col}{first_data_row}))=0'],
            fill=error_fill,
            stopIfTrue=False,
        ),
    )

    # Conditional formatting: highlight invalid ID (filled but doesn't match pattern)
    ws.conditional_formatting.add(
        id_range,
        FormulaRule(
            formula=[
                f'AND(LEN(TRIM(C{first_data_row}))>0,NOT(AND(LEN(C{first_data_row})=10,LEFT(C{first_data_row},4)="PRD-",ISNUMBER(VALUE(RIGHT(C{first_data_row},6))))))'
            ],
            fill=warn_fill,
            stopIfTrue=False,
        ),
    )

    # Borders for a larger input area (keeps it looking like a sheet)
    for row in range(first_data_row, 51):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is None:
                cell.value = ""
            cell.alignment = body_alignment
            cell.border = border

    ws.sheet_view.showGridLines = False

    for col_idx in range(1, len(headers) + 1):
        ws.cell(row=header_row, column=col_idx).border = border

    ws.print_title_rows = f"{title_row}:{header_row}"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    ws.auto_filter.ref = table_ref

    ws.column_dimensions["C"].bestFit = True
    ws.column_dimensions["C"].hidden = False

    for col in range(1, len(headers) + 1):
        ws.cell(row=title_row, column=col).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=info_row, column=col).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    try:
        wb.save(path)
    finally:
        wb.close()


def main() -> None:
    generate_xlsx(Path("artifacts/produtos_entrada_template.xlsx"))


if __name__ == "__main__":
    main()

