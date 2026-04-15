from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SUMMARY_HEADERS = [
    "Fonte",
    "ID do produto",
    "Nome informado",
    "URL",
    "CEP",
    "Nome do produto",
    "Valor do frete",
    "Moeda",
    "Tipo do frete",
    "Prazo de entrega",
    "Modo de entrega",
    "Quantidade de opções",
    "Opções de frete",
]

OPTIONS_HEADERS = [
    "Fonte",
    "ID do produto",
    "Nome informado",
    "URL",
    "CEP",
    "Índice da opção",
    "Prazo de entrega",
    "Modo de entrega",
    "Valor",
    "Tipo do frete",
    "Texto do valor",
]


def _normalize_options(value: Any) -> list[dict[str, Any]]:
    if isinstance(value, list):
        out: list[dict[str, Any]] = []
        for item in value:
            if isinstance(item, dict):
                out.append(dict(item))
        return out
    return []


def _format_price_kind(value: Any) -> str:
    kind = str(value or "").upper()
    if kind == "FREE":
        return "Grátis"
    if kind == "PAID":
        return "Pago"
    return "Indisponível"


def _format_option_line(index: int, option: dict[str, Any]) -> str:
    price_kind = _format_price_kind(option.get("price_kind"))
    price_text = option.get("price_text")
    price = option.get("price")
    if price_text:
        value_text = str(price_text)
    elif price is None:
        value_text = "Valor indisponível"
    else:
        value_text = f"R$ {price}"

    parts = [
        f"{index}.",
        value_text,
        f"Tipo: {price_kind}",
    ]
    if option.get("delivery_time_text"):
        parts.append(f"Prazo: {option.get('delivery_time_text')}")
    if option.get("delivery_mode"):
        parts.append(f"Modo: {option.get('delivery_mode')}")
    return " | ".join(parts)


def _append_headers(ws, headers: list[str]) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _autosize_columns(ws) -> None:
    for column_cells in ws.columns:
        values = []
        for cell in column_cells:
            if cell.value is None:
                continue
            values.append(str(cell.value))
        if not values:
            continue
        width = min(max(len(v) for v in values) + 2, 60)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def build_results_workbook(jobs: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Resultados"

    options_ws = wb.create_sheet("Frete Opções")
    _append_headers(summary_ws, SUMMARY_HEADERS)
    _append_headers(options_ws, OPTIONS_HEADERS)

    summary_rows = 0
    options_rows = 0

    for job in jobs:
        result = job.get("result") if isinstance(job, dict) else None
        result_dict = result if isinstance(result, dict) else {}
        freight = result_dict.get("freight") if isinstance(result_dict.get("freight"), dict) else {}
        options = _normalize_options(freight.get("options"))

        summary_row = {
            "Fonte": result_dict.get("source") or job.get("group") or "",
            "ID do produto": job.get("product_id"),
            "Nome informado": job.get("input_product_name"),
            "URL": job.get("url"),
            "CEP": job.get("cep"),
            "Nome do produto": result_dict.get("product_name"),
            "Valor do frete": freight.get("price"),
            "Moeda": freight.get("currency"),
            "Tipo do frete": _format_price_kind(freight.get("price_kind")),
            "Prazo de entrega": freight.get("delivery_time_text"),
            "Modo de entrega": freight.get("delivery_mode"),
            "Quantidade de opções": len(options),
            "Opções de frete": "\n".join(
                _format_option_line(idx, option) for idx, option in enumerate(options, start=1)
            ),
        }

        summary_rows += 1
        summary_ws.append([summary_row.get(h) for h in SUMMARY_HEADERS])
        summary_ws.cell(row=summary_rows + 1, column=len(SUMMARY_HEADERS)).alignment = Alignment(
            horizontal="left",
            vertical="top",
            wrap_text=True,
        )

        if not options:
            options = [{}]
        for idx, option in enumerate(options, start=1):
            options_rows += 1
            options_ws.append(
                [
                    result_dict.get("source") or job.get("group") or "",
                    job.get("product_id"),
                    job.get("input_product_name"),
                    job.get("url"),
                    job.get("cep"),
                    idx,
                    option.get("delivery_time_text"),
                    option.get("delivery_mode"),
                    option.get("price"),
                    _format_price_kind(option.get("price_kind")),
                    option.get("price_text"),
                ]
            )

    summary_ws.freeze_panes = "A2"
    options_ws.freeze_panes = "A2"
    summary_ws.auto_filter.ref = f"A1:{get_column_letter(len(SUMMARY_HEADERS))}{max(summary_rows, 1) + 1}"
    options_ws.auto_filter.ref = f"A1:{get_column_letter(len(OPTIONS_HEADERS))}{max(options_rows, 1) + 1}"
    _autosize_columns(summary_ws)
    _autosize_columns(options_ws)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
