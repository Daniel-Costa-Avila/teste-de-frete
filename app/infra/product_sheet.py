from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass

from openpyxl import load_workbook


@dataclass(frozen=True)
class ProductInputRow:
    group: str
    product_name: str
    product_id: str
    url: str
    ceps: tuple[str, ...]


EXPECTED_HEADERS_V1 = ["nome do produto", "id do produto", "link do produto"]
EXPECTED_HEADERS_V2 = ["grupo", "nome do produto", "id do produto", "link do produto", "ceps para testar"]

_CEP_RE = re.compile(r"\b\d{5}-?\d{3}\b")


def _norm_header(value: object) -> str:
    return str(value or "").strip().lower()


def _parse_ceps(value: object) -> tuple[str, ...]:
    raw = str(value or "").strip()
    if not raw:
        return ()

    matches = [m.group(0) for m in _CEP_RE.finditer(raw)]

    normalized: list[str] = []
    for m in matches:
        digits = re.sub(r"\D", "", m)
        if len(digits) == 8:
            normalized.append(f"{digits[:5]}-{digits[5:]}")

    seen: set[str] = set()
    out: list[str] = []
    for c in normalized:
        if c in seen:
            continue
        seen.add(c)
        out.append(c)
    return tuple(out)


def parse_products_csv(data: bytes) -> list[ProductInputRow]:
    text = data.decode("utf-8-sig", errors="replace")
    f = io.StringIO(text)
    reader = csv.reader(f, delimiter=";")

    rows = list(reader)
    if not rows:
        return []

    header = [_norm_header(c) for c in rows[0]]
    header3 = header[:3]
    header5 = header[:5]

    v2 = header5 == EXPECTED_HEADERS_V2
    v1 = header3 == EXPECTED_HEADERS_V1
    if not v1 and not v2:
        raise ValueError(
            "Cabeçalhos inválidos no CSV (esperado: "
            "Nome do produto;ID do produto;Link do produto "
            "ou Grupo;Nome do produto;ID do produto;Link do produto;CEPs para testar)."
        )

    out: list[ProductInputRow] = []
    for r in rows[1:]:
        if not r or all((c or "").strip() == "" for c in r):
            continue

        if v2:
            group = (r[0] if len(r) > 0 else "").strip()
            name = (r[1] if len(r) > 1 else "").strip()
            pid = (r[2] if len(r) > 2 else "").strip()
            url = (r[3] if len(r) > 3 else "").strip()
            ceps_raw = (r[4] if len(r) > 4 else "").strip()
        else:
            group = ""
            name = (r[0] if len(r) > 0 else "").strip()
            pid = (r[1] if len(r) > 1 else "").strip()
            url = (r[2] if len(r) > 2 else "").strip()
            ceps_raw = ""

        if not name and not pid and not url and not group and not ceps_raw:
            continue
        if not url:
            continue

        out.append(
            ProductInputRow(
                group=group,
                product_name=name,
                product_id=pid,
                url=url,
                ceps=_parse_ceps(ceps_raw),
            )
        )

    return out


def parse_products_xlsx(data: bytes) -> list[ProductInputRow]:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active

    header_row_idx: int | None = None
    header_map: dict[str, int] = {}
    for i in range(1, 31):
        values = [ws.cell(row=i, column=j).value for j in range(1, 9)]
        normalized = [_norm_header(v) for v in values]
        if "link do produto" not in normalized:
            continue

        header_map = {h: idx for idx, h in enumerate(normalized) if h}
        if "nome do produto" in header_map and "id do produto" in header_map and "link do produto" in header_map:
            header_row_idx = i
            break

    if header_row_idx is None:
        raise ValueError(
            "Cabeçalhos não encontrados no XLSX (esperado: Nome do produto / ID do produto / Link do produto "
            "ou Grupo / ... / CEPs para testar)."
        )

    def _get(row: tuple[object, ...], col_name: str) -> str:
        idx = header_map.get(col_name)
        if idx is None or idx >= len(row):
            return ""
        v = row[idx]
        return str(v).strip() if v is not None else ""

    out: list[ProductInputRow] = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not row or all((c is None or str(c).strip() == "") for c in row):
            continue

        group = _get(row, "grupo")
        name = _get(row, "nome do produto")
        pid = _get(row, "id do produto")
        url = _get(row, "link do produto")
        ceps_raw = _get(row, "ceps para testar")

        if not url:
            continue

        out.append(
            ProductInputRow(
                group=group,
                product_name=name,
                product_id=pid,
                url=url,
                ceps=_parse_ceps(ceps_raw),
            )
        )

    return out


def parse_products_file(filename: str, data: bytes) -> list[ProductInputRow]:
    name = (filename or "").lower().strip()
    if name.endswith(".csv"):
        return parse_products_csv(data)
    if name.endswith(".xlsx"):
        return parse_products_xlsx(data)
    raise ValueError("Formato de arquivo não suportado. Envie .xlsx ou .csv.")

